# Duplicate Directory Finder
import time
import os
import sys
from openpyxl import Workbook
import ddf
from operator import itemgetter

# Get the two working directories from the user
if len(sys.argv) > 3:
    print('Please put folder names with spaces inside of double quotes')
    sys.exit()

if len(sys.argv) == 3:
    if os.path.isdir(sys.argv[1]):
        target_1 = sys.argv[1]
    else:
        print('The first directory is not valid')
        sys.exit()
    if os.path.isdir(sys.argv[2]):
        target_2 = sys.argv[2]
    else:
        print('The second directory is not valid')
        sys.exit()
else:
    print('Please provide two valid directories after the program name.')
    sys.exit()

# Set up the Duplicate Directory Objects
finder_1 = ddf.DuplicateDirectoryFinder(target_1)
finder_2 = ddf.DuplicateDirectoryFinder(target_2)

# Set up the timestamps 
start = time.time()
nothing, nothing, nothing, nothing = finder_1.find_duplicates()
nothing, nothing, nothing, nothing = finder_2.find_duplicates()
end = time.time()

# Create a unique string name for the time when the program finished
end_string = time.strftime("%Y-%m-%d %H_%M_%S", time.gmtime(end))

# Print the timestamps to the user
print('The program ran for', (end - start),' seconds.')

# Get all of the non-zero size duplicate folders in both directories
directories_1 = []
directories_2 = []
for a in range (0, len(finder_1.directory_hashes)):
    dir1, dir2, num_dirs, num_files, dir_size =  finder_1.directory_hashes[a]
    if dir_size != 0:
        directories_1.append((dir1, dir2, num_dirs, num_files, dir_size))

for a in range (0, len(finder_2.directory_hashes)):
    dir1, dir2, num_dirs, num_files, dir_size =  finder_2.directory_hashes[a]
    if dir_size != 0:
        directories_2.append((dir1, dir2, num_dirs, num_files, dir_size))

# Compile the list of duplicates from both directories
dupes = []
for a in range (0, len(directories_1)):
    for b in range (0, len(directories_2)):
        if directories_1[a][0] == directories_2[b][0]:
            dupes.append(directories_1[a] + directories_2[b])

# Sort the list of duplicate directories by size
dupes = sorted(dupes, key=itemgetter(4), reverse = True)

# Save entire array to Excel file using OpenPyXL
print('Saving Results to file')
print()
wb = Workbook()
ws = wb.active

# Set up header info in Excel file
ws['A1'] = 'Hash Value'
ws['B1'] = 'Directory Name 1'
ws['C1'] = 'Directory Name 2'
ws['D1'] = 'Number of Subdirectories'
ws['E1'] = 'Number of Files'
ws['F1'] = 'Size of Directory'

# Loop through the list of hashes and save the four parts of the tuple
for a in range (0, len(dupes)):
    ws['A' + str(a + 2)] = dupes[a][0]
    ws['B' + str(a + 2)] = dupes[a][1]
    ws['C' + str(a + 2)] = dupes[a][6]
    ws['D' + str(a + 2)] = dupes[a][2]
    ws['E' + str(a + 2)] = dupes[a][3]
    ws['F' + str(a + 2)] = dupes[a][4]

# Save the file in the current directory
wb.save('Duplicate Folders ' + end_string + '.xlsx')

print('Results saved as Duplicate Folders ' + end_string + '.xlsx')
print()
junk = input('Press Enter to exit the program')

