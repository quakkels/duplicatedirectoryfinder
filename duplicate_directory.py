# Duplicate Directory Finder

import time
import os
import sys
from openpyxl import Workbook
import ddf
from operator import itemgetter

# Get the current working directory (if supplied by user)
if len(sys.argv) > 2:
    print('Please put folder names with spaces inside of double quotes')
    sys.exit()

if len(sys.argv) == 2:
    if os.path.isdir(sys.argv[1]):
        start_directory = sys.argv[1]
    else:
        print('That was not a valid directory')
        sys.exit()
else:
    print("Please provide a valid directory.")
    sys.exit()

# Set up the Duplicate Directory Object
finder = ddf.DuplicateDirectoryFinder(start_directory)

# Set up the timestamps 
start = time.time()
dir_count, file_count, folder_size, nothing = finder.find_duplicates()
end = time.time()

# Create a unique string name for the time when the program finished
end_string = time.strftime("%Y-%m-%d %H_%M_%S", time.gmtime(end))

# Print the timestamps to the user
print('The program ran for', (end - start),' seconds.')

# Save entire array to Excel file using OpenPyXL
print('Saving Results to file')
print()
wb = Workbook()
ws = wb.active

# Set up header info in Excel file
ws['A1'] = 'Hash Value'
ws['B1'] = 'Directory Name'
ws['C1'] = 'Number of Subdirectories'
ws['D1'] = 'Number of Files'
ws['E1'] = 'Size of Directory'


# Loop through the list of hashes and save the four parts of the tuple
for a in range (0, len(finder.directory_hashes)):
    hash, dir_name, num_sub_dirs, num_files, dir_size = finder.directory_hashes[a]
    ws['A' + str(a + 2)] = hash
    ws['B' + str(a + 2)] = dir_name
    ws['C' + str(a + 2)] = num_sub_dirs
    ws['D' + str(a + 2)] = num_files
    ws['E' + str(a + 2)] = dir_size

# Save the file in the current directory
wb.save('Folder Hashes ' + end_string + '.xlsx')

# Save just the duplicate files to another to Excel file
wb = Workbook()
ws = wb.active
ws['A1'] = 'Directory 1'
ws['B1'] = 'Directory 2'
ws['C1'] = 'Number of Subdirectories'
ws['D1'] = 'Number of Files'
ws['E1'] = 'Size of Directory'

dupes = []
# Get all of the non-zero size duplicate folders
for a in range (0, len(finder.found_duplicates)):
    dir1, dir2, num_dirs, num_files, dir_size =  finder.found_duplicates[a]
    if num_files != 0 and dir_size != 0:
        dupes.append([dir1, dir2, num_dirs, num_files, dir_size])

dupes = sorted(dupes, key=itemgetter(4), reverse = True)

# Save the duplicate folders to the Excel file
for a in range (0, len(dupes)):
    ws['A' + str(a + 2)] = dupes[a][0]
    ws['B' + str(a + 2)] = dupes[a][1]
    ws['C' + str(a + 2)] = dupes[a][2]
    ws['D' + str(a + 2)] = dupes[a][3]
    ws['E' + str(a + 2)] = dupes[a][4]
wb.save('Duplicate Folders ' + end_string + '.xlsx')

print('Results saved as Folder Hashes ' + end_string + 
      '.xlsx and Duplicate Folders ' + end_string + '.xlsx')
print()
junk = input('Press Enter to exit the program')

